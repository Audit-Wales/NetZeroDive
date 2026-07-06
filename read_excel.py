import openpyxl
wb = openpyxl.load_workbook(r'c:\Users\vicky.jones\Git Hub\DIVE - Net Zero\Welsh_Public_Sector_Net_Zero_2025_with_Insights (1).xlsx', data_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'=== Sheet: {name} ({ws.max_row} rows x {ws.max_column} cols) ===')
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        print(row)
    print()
